import tkinter as tk
import openpyxl
def get_user_name():
    user_name = entry_name.get()
    if len(user_name) > 10:
        print("Please enter a name with less than 10 characters.")
        return get_user_name()
    else:
        return user_name

def get_user_age():
    try:
        user_age = int(entry_age.get())
        return user_age
    except ValueError:
        print("Please enter a valid age.")
        return get_user_age()

def get_user_gender():
    try:
        user_gender = int(entry_gender.get())
        if user_gender == 1:
            user_gender = "male"
        elif user_gender == 2:
            user_gender = "female"
        else:
            print("Please enter a valid gender.")
            return get_user_gender()
        return user_gender
    except ValueError:
        print("Please enter a valid gender.")
        return get_user_gender()

root = tk.Tk()
root.title("User Profile")

label_name = tk.Label(root, text="What is your character's name?")
label_name.pack()
entry_name = tk.Entry(root)
entry_name.pack()

label_age = tk.Label(root, text="What is your age?")
label_age.pack()
entry_age = tk.Entry(root)
entry_age.pack()

label_gender = tk.Label(root, text="What is your gender? (1 for male, 2 for female)")
label_gender.pack()
entry_gender = tk.Entry(root)
entry_gender.pack()

button_submit = tk.Button(root, text="Submit", command=lambda: save_user_profile(get_user_name(), get_user_age(), get_user_gender()))
button_submit.pack()

root.mainloop()
def save_user_profile(user_name, user_age, user_gender):
    userprofile = openpyxl.load_workbook("userprofile.xlsx")
    profilesheet = userprofile.active

    profilesheet[f'A{profilesheet.max_row + 1}'] = user_name
    profilesheet[f'B{profilesheet.max_row}'] = user_age
    profilesheet[f'C{profilesheet.max_row}'] = user_gender

    userprofile.save("userprofile.xlsx")

    print("Data saved")

