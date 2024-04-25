import openpyxl

userprofile = openpyxl.load_workbook("userprofile.xlsx")
profilesheet = userprofile.active

def get_user_name():
    while True:
        user_name = input("What is your character's name? ")
        
        if len(user_name) > 10:
            print("Please enter a name with less than 10 characters.")
            continue
        else:
            return user_name

def get_user_age():
    while True:
        try:
            user_age = int(input("What is your age? "))
            return user_age
        except ValueError:
            print("Please enter a valid age.")

def get_user_gender():
    while True:
        try:
            user_gender = int(input("What is your gender? (1 for male, 2 for female) "))
            if user_gender == 1:
                user_gender = "male"
            elif user_gender == 2:
                user_gender = "female"
            else:
                print("Please enter a valid gender.")
                continue
            return user_gender
        except ValueError:
            print("Please enter a valid gender.")

user_name = get_user_name()
user_age = get_user_age()
user_gender = get_user_gender()

userprofile = openpyxl.load_workbook("userprofile.xlsx")
profilesheet = userprofile.active


profilesheet[f'A{profilesheet.max_row + 1}'] = user_name
profilesheet[f'B{profilesheet.max_row}'] = user_age
profilesheet[f'C{profilesheet.max_row}'] = user_gender


userprofile.save("userprofile.xlsx")

print("Data saved")




 
